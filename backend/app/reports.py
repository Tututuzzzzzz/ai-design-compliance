"""Batch report export — CSV and Excel."""

from __future__ import annotations

import csv
import datetime
import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import settings

from .pipeline import i18n
from .pipeline.i18n import DEFAULT_LANG, Lang

log = logging.getLogger(__name__)

#: Column order. Headers are looked up per language at export time.
COLUMN_KEYS = [
    "filename",
    "scanned",
    "verdict",
    "confidence",
    "niche",
    "sub_niche",
    "style",
    "motifs",
    "risk_categories",
    "finding_count",
    "top_severity",
    "rights_holders",
    "trademark_matches",
    "regions",
    "remediation",
    "markets",
    "platforms",
    "ocr_text",
    "source",
    "source_ref",
    "reasoning",
    "error",
]


def columns(lang: Lang) -> list[tuple[str, str]]:
    return [(key, i18n.export(key, lang)) for key in COLUMN_KEYS]

_SEV_ORDER = {"low": 0, "medium": 1, "high": 2, "critical": 3}

FILLS = {
    "SAFE": PatternFill("solid", fgColor="D1FAE5"),
    "RISKY": PatternFill("solid", fgColor="FEF3C7"),
    "BLOCKED": PatternFill("solid", fgColor="FEE2E2"),
}


def flatten(design: dict[str, Any], lang: Lang = DEFAULT_LANG) -> dict[str, Any]:
    report = design.get("report") or {}
    niche = report.get("niche") or {}
    findings = report.get("findings") or []
    hits = report.get("trademark_hits") or []
    meta = report.get("metadata") or {}

    top_sev = ""
    if findings:
        top_sev = max(findings, key=lambda f: _SEV_ORDER.get(f.get("severity", "low"), 0))[
            "severity"
        ]

    verdict_value = report.get("verdict") or design.get("verdict")
    return {
        "filename": design.get("filename", ""),
        "scanned": _scanned(design.get("created_at")),
        "verdict": (
            i18n.verdict(verdict_value, lang) if verdict_value else i18n.export("failed", lang)
        ),
        "confidence": report.get("confidence", 0),
        "niche": niche.get("primary", ""),
        "sub_niche": niche.get("sub_niche") or "",
        "style": ", ".join(niche.get("style") or []),
        "motifs": ", ".join(niche.get("motifs") or []),
        "risk_categories": ", ".join(
            sorted({i18n.category(f.get("category", ""), lang) for f in findings})
        ),
        "finding_count": len(findings),
        "top_severity": i18n.severity_inline(top_sev, lang) if top_sev else "",
        "rights_holders": ", ".join(
            sorted({f["rights_holder"] for f in findings if f.get("rights_holder")})
        ),
        "trademark_matches": " | ".join(
            f"{h.get('mark_text')} ({h.get('similarity')}%"
            + (f", #{h['serial_number']}" if h.get("serial_number") else "")
            + ")"
            for h in hits
        ),
        "regions": " | ".join(_region(f) for f in findings if f.get("bbox") or f.get("location_hint")),
        "remediation": " | ".join(f.get("remediation", "") for f in findings),
        "markets": ", ".join(meta.get("markets") or []),
        "platforms": ", ".join(meta.get("platforms") or []),
        "ocr_text": (report.get("ocr_text") or "").replace("\n", " / "),
        "source": design.get("source", ""),
        "source_ref": design.get("source_ref") or "",
        "reasoning": report.get("reasoning", ""),
        "error": design.get("error") or report.get("error") or "",
    }


def _scanned(epoch: Any) -> str:
    """Local wall-clock stamp for the row, or blank if the row predates the column.

    Written as a plain string rather than a real datetime: the CSV and the XLSX
    share this dict, and a spreadsheet locale-guessing the format is worse than
    an unambiguous ISO-ish stamp.
    """
    if not epoch:
        return ""
    try:
        return datetime.datetime.fromtimestamp(float(epoch)).strftime("%Y-%m-%d %H:%M")
    except (TypeError, ValueError, OSError):
        return ""


def _region(f: dict[str, Any]) -> str:
    box = f.get("bbox")
    hint = f.get("location_hint") or f.get("title", "")
    if not box:
        return hint
    return (
        f"{hint} [x={box['x']:.2f} y={box['y']:.2f} w={box['w']:.2f} h={box['h']:.2f}]"
    )


def to_csv(designs: list[dict[str, Any]], lang: Lang = DEFAULT_LANG) -> bytes:
    cols = columns(lang)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[k for k, _ in cols], extrasaction="ignore")
    writer.writerow({k: label for k, label in cols})
    for d in designs:
        writer.writerow(flatten(d, lang))
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 correctly


def to_xlsx(
    designs: list[dict[str, Any]], stats: dict[str, int], lang: Lang = DEFAULT_LANG
) -> bytes:
    cols = columns(lang)
    wb = Workbook()

    ws = wb.active
    ws.title = i18n.export("sheet.summary", lang)
    ws.append([i18n.export("metric", lang), i18n.export("count", lang)])
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    total = sum(stats.values())
    for key in ("SAFE", "RISKY", "BLOCKED", "FAILED"):
        label = i18n.verdict(key, lang) if key in FILLS else i18n.export("failed", lang)
        ws.append([label, stats.get(key, 0)])
        if key in FILLS:
            ws.cell(row=ws.max_row, column=1).fill = FILLS[key]
    ws.append([i18n.export("total", lang), total])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10

    det = wb.create_sheet(i18n.export("sheet.designs", lang))
    det.append([label for _, label in cols])
    for cell in det[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    det.freeze_panes = "A2"

    for d in designs:
        row = flatten(d, lang)
        det.append([row[k] for k, _ in cols])
        # Colour by the raw verdict, not the translated label.
        raw_verdict = (d.get("report") or {}).get("verdict") or d.get("verdict")
        fill = FILLS.get(str(raw_verdict))
        if fill:
            det.cell(row=det.max_row, column=2).fill = fill

    det.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{det.max_row}"
    widths = {"filename": 34, "reasoning": 70, "remediation": 60, "ocr_text": 40}
    for i, (key, _) in enumerate(cols, start=1):
        det.column_dimensions[get_column_letter(i)].width = widths.get(key, 18)

    fnd = wb.create_sheet(i18n.export("sheet.findings", lang))
    fnd.append([
        i18n.export("filename", lang),
        i18n.export("category", lang),
        i18n.export("severity", lang),
        i18n.export("confidence", lang),
        i18n.export("title", lang),
        i18n.export("rights_holder", lang),
        i18n.export("matched_text", lang),
        i18n.export("location", lang),
        i18n.export("evidence", lang),
        i18n.export("remediation", lang),
    ])
    for cell in fnd[1]:
        cell.font = Font(bold=True)
    for d in designs:
        report = d.get("report") or {}
        for f in report.get("findings") or []:
            fnd.append([
                d.get("filename", ""),
                i18n.category(f.get("category", ""), lang),
                i18n.severity_inline(f.get("severity", ""), lang),
                f"{float(f.get('confidence') or 0):.0%}",
                f.get("title", ""),
                f.get("rights_holder") or "",
                f.get("matched_text") or "",
                _region(f),
                " ; ".join(
                    (e.get("reference_id") or e.get("detail") or "")
                    for e in (f.get("evidence") or [])
                ),
                f.get("remediation", ""),
            ])
    for i, w in enumerate([30, 24, 12, 12, 46, 24, 24, 34, 44, 56], start=1):
        fnd.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------
# Review-sheet export — the organiser's manifest layout, with the design
# thumbnail embedded in the cell.
# --------------------------------------------------------------------------

#: Column order and header text copied verbatim from the organiser's workbook so
#: the file is drop-in comparable against their answer key. The `expected_*`
#: prefix is theirs and is kept even though these cells hold *our* verdict — a
#: renamed header would stop the two files lining up column-for-column, which is
#: the only reason to emit this layout at all.
SHEET_COLUMNS = [
    "no",
    "design",
    "target_market",
    "platform",
    "expected_niche",
    "expected_sub_niche",
    "expected_style",
    "expected_motifs",
    "expected_verdict",
    "expected_violation_type",
    "expected_violation_detail",
    "expected_confidence",
    "notes",
]

#: Our seven categories collapsed back onto the sheet's coarser vocabulary.
#: This is the inverse of `data/import_sheet.CATEGORY_MAP`; when a row has
#: several findings the most specific label wins, in this order.
_SHEET_VIOLATION_LABEL = {
    "copyrighted_character": "Character",
    "brand_logo": "Logo",
    "trademarked_phrase": "Text/Band Name",
    "public_figure": "Celebrity Likeness",
    "copyrighted_artwork": "Artwork/Copyright",
    "licensed_font": "Font",
    "prohibited_content": "Sensitive/Prohibited",
}
_SHEET_LABEL_ORDER = [
    "copyrighted_character",
    "public_figure",
    "brand_logo",
    "trademarked_phrase",
    "copyrighted_artwork",
    "prohibited_content",
    "licensed_font",
]

#: Pixel size of the embedded thumbnail. Small on purpose: 500 designs at
#: preview resolution would produce a workbook nobody can open.
SHEET_THUMB_PX = 96


def _sheet_violation_type(findings: list[dict[str, Any]]) -> str:
    present = {f.get("category") for f in findings}
    for category in _SHEET_LABEL_ORDER:
        if category in present:
            return _SHEET_VIOLATION_LABEL[category]
    return ""


def _thumbnail(path: Path, px: int = SHEET_THUMB_PX) -> io.BytesIO | None:
    """Square-ish RGB thumbnail of a preview, ready to embed.

    Flattened, not passed through as-is: the preview keeps the design's
    transparency, and Excel paints cells white — a white wordmark on a
    transparent background would be invisible in the very report meant to
    document it.
    """
    from PIL import Image  # noqa: PLC0415

    from .pipeline import loader  # noqa: PLC0415

    try:
        with Image.open(path) as img:
            img.load()
            flat = loader.flatten(img)
        flat.thumbnail((px, px), Image.LANCZOS)
        buf = io.BytesIO()
        flat.save(buf, format="PNG", optimize=True)
        buf.seek(0)
        return buf
    except Exception as exc:  # pragma: no cover - a bad preview must not kill the export
        log.warning("Could not build thumbnail for %s: %s", path, exc)
        return None


def _preview_path(design: dict[str, Any]) -> Path | None:
    """Locate the preview PNG on disk from the URL stored in the report."""
    url = (design.get("report") or {}).get("preview_url") or ""
    name = url.rsplit("/", 1)[-1]
    if not name:
        return None
    path = settings.renders_dir / name
    return path if path.exists() else None


def to_sheet_xlsx(
    designs: list[dict[str, Any]], lang: Lang = DEFAULT_LANG
) -> bytes:
    """One row per design in the organiser's column layout, image included."""
    wb = Workbook()
    ws = wb.active
    ws.title = i18n.export("sheet.review", lang)

    ws.append(SHEET_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"

    # Keep every buffer referenced until save(); openpyxl reads them lazily.
    held: list[io.BytesIO] = []

    for i, d in enumerate(designs, start=1):
        report = d.get("report") or {}
        niche = report.get("niche") or {}
        findings = report.get("findings") or []
        meta = report.get("metadata") or {}

        ws.append([
            i,
            "",  # filled by the embedded image below
            ", ".join(meta.get("markets") or []),
            ", ".join(meta.get("platforms") or []),
            niche.get("primary", ""),
            niche.get("sub_niche") or "",
            ", ".join(niche.get("style") or []),
            ", ".join(niche.get("motifs") or []),
            report.get("verdict") or d.get("verdict") or "",
            _sheet_violation_type(findings),
            " | ".join(f.get("description", "") for f in findings),
            report.get("confidence", 0),
            report.get("reasoning", ""),
        ])
        row = ws.max_row

        # The filename no longer has a column of its own — the layout is fixed at
        # the organiser's 13 — so keep it reachable as a comment on the image
        # cell. Losing which file a row came from would make the sheet unusable
        # for anything but eyeballing.
        filename = d.get("filename", "")
        if filename:
            ws.cell(row=row, column=2).comment = Comment(filename, "compliance")

        raw_verdict = str(report.get("verdict") or d.get("verdict") or "")
        fill = FILLS.get(raw_verdict)
        if fill:
            ws.cell(row=row, column=9).fill = fill

        ws.cell(row=row, column=11).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=13).alignment = Alignment(vertical="top", wrap_text=True)

        preview = _preview_path(d)
        thumb = _thumbnail(preview) if preview else None
        if thumb is not None:
            held.append(thumb)
            image = XLImage(thumb)
            # Anchor to the cell; openpyxl sizes from the PNG's own pixels.
            ws.add_image(image, f"B{row}")
            ws.row_dimensions[row].height = SHEET_THUMB_PX * 0.78
        else:
            ws.cell(row=row, column=2).value = i18n.export("noPreview", lang)

    widths = {
        "no": 5,
        "design": 15,
        "target_market": 13,
        "platform": 15,
        "expected_niche": 18,
        "expected_sub_niche": 20,
        "expected_style": 16,
        "expected_motifs": 34,
        "expected_verdict": 15,
        "expected_violation_type": 20,
        "expected_violation_detail": 60,
        "expected_confidence": 12,
        "notes": 72,
    }
    for i, key in enumerate(SHEET_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[key]

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET_COLUMNS))}{ws.max_row}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
